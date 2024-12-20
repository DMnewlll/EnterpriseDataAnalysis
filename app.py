import csv
import io
import os
import re
import pandas as pd
import requests
from flask import Flask, request, redirect, url_for, render_template, flash
from flask import Flask, json
from impala.dbapi import connect
from hdfs import InsecureClient
from openpyxl.reader.excel import load_workbook
from collections import OrderedDict
from sklearn.cluster import KMeans
import matplotlib.pyplot as plt
import seaborn as sns
from flask_cors import CORS
from decimal import Decimal
from flask import Flask, jsonify

app = Flask(__name__)

HOST = "121.36.46.125"
HDFS_URL = 'http://121.36.46.125:9870'
HDFS_NAMENODE_HOST = 'http://121.36.46.125:9870/webhdfs/v1'
HDFS_USER = 'root'
HDFS_UPLOAD_DIR = '/user/data/'
HIVE_PORT = '10000'


areas1 = {'4501': '南宁', '4502': '柳州', '4503': '桂林', '4504': '梧州', '4505': '北海', '4506': '防城港',
          '4507': '钦州', '4508': '贵港', '4509': '玉林', '4510': '百色', '4511': '贺州', '4512': '河池',
          '4513': '来宾',
          '4514': '崇左'}
areas = {'南宁': '4501', '柳州': '4502', '桂林': '4503', '梧州': '4504', '北海': '4505', '防城港': '4506',
         '钦州': '4507',
         '贵港': '4508', '玉林': '4509', '百色': '4510', '贺州': '4511', '河池': '4512', '来宾': '4513',
         '崇左': '4514'}


# 主页面路由
@app.route("/", methods=["POST", "GET"])
def index():
    return render_template('main.html')

# 可视化表路由
@app.route("/upload1")
def upload1():
    print("Rendering demo_1.html")  # 调试输出，检查是否进入函数
    return render_template('demo_1.html')  # 渲染 demo_1.html 页面


#显示上传文件页面路由
@app.route("/shangchuan")
def shangchuan():
    print("Rendering upload.html")  # 调试输出，检查是否进入函数
    return render_template('upload.html')  # 渲染 upload.html 页面


# 上传文件到HDFS，从HDFS获取文件进行清洗后导入hive数据仓库
@app.route("/upload", methods=["POST", "GET"])
def upload():
    if request.method == 'POST':
        # 1.处理上传的文件
        # 判断请求中是否包含文件
        if 'file' not in request.files:
            return render_template('upload.html', error="没有文件部分")

        file = request.files['file']

        # 判断有没有选择文件
        if file.filename == '':
            return render_template('upload.html', error="没有选择文件")

        try:
            # 获取上传的文件的内容
            file_content = file.read()
            workbook = load_workbook(io.BytesIO(file_content), data_only=True)
            # 第一个工作表
            sheet = workbook.active
            # 将文件转化为csv文件
            # 准备CSV文件名
            csv_filename = os.path.join('高企源数据.csv')
            # 写入CSV文件
            with open(csv_filename, mode='w', newline='', encoding='utf-8') as csv_file:
                writer = csv.writer(csv_file)
                header = [cell.value for cell in sheet[1]]  # 第一行作为表头
                writer.writerow(header)
                # 写入数据行
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    writer.writerow(row)

            # 3.将源数据上传到HDFS
            hdfs_path = os.path.join(HDFS_UPLOAD_DIR, file.filename)
            url = f"{HDFS_NAMENODE_HOST}{hdfs_path}?op=CREATE&user.name={HDFS_USER}&overwrite=true"
            headers = {'Content-Type': 'application/octet-stream'}
            response = requests.put(url, data=file_content, headers=headers)
            if response.status_code == 201:
                print("上传源数据成功")
            else:
                print("上传源数据失败")

            # 4.进行数据清洗
            df = pd.read_csv(csv_filename)
            df = df.drop([0, 1], axis=0)
            df = df.reset_index(drop=True)
            df['行政区划代码'] = df['行政区划代码'].str[:4]
            df = df[df['数据状态'] != 'dis']

            columns_to_drop = ['序号', '所属地域', '数据年份', '撤消原因',
                               '申请撤消状态', '管理员审核状态', '错误数量', '警告数量', '数据提交时间',
                               '数据提交状态', '地址', '是否填写国家统计局一套表',
                               '企业统一社会信用代码或组织机构代码', '法人性质', '邮政编码',
                               '企业注册地址', '企业注册地是否在国家高新区内',
                               '企业主要生产经营活动是否在国家高新区内', '已进区企业被批准入区时间',
                               '出生年份', '企业是否进入国家自主创新示范区',
                               '如是，请选择所在自创区名称', '填报时间', '企业隶属关系',
                               '主要业务活动或主要产品1', '主要业务活动或主要产品2',
                               '主要业务活动或主要产品3', '注册时间',
                               '执行会计标准类别', '企业执行会计准则情况',
                               '登记注册类型', '主要外资来源国别或地区代码',
                               '主要外资出资比例', '控股情况', '企业集团情况',
                               'QB09如为2，请填直接上级法人单位组织机构代码或上级法人统一社会信用代码',
                               '是否为经过认定的高新技术企业', '高新技术企业认定证书编号',
                               '企业被认定为高新技术企业的时间', '与科技企业孵化器关系',
                               '入孵时间', '毕业时间',
                               '股票代码',
                               '如qb15_3为否，请填写上市（挂牌）公司主体名称',
                               '企业所属技术领域',
                               '企业核心技术所属国家重点支持的高新技术领域',
                               '进出口总额', '其中：出口总额',
                               '其中：高新技术产品出口', '技术服务出口', '营业成本',
                               '其中：研发、试验检验费', '税金及附加', '销售费用',
                               '管理费用', '其中：技术（研究）开发费',
                               '其中：支付科研人员的工资及福利费', '财务费用',
                               '资产减值损失', '公允价值变动收益（损失以“-”号记）',
                               '投资收益（损失以“-”号记）', '资产处置收益（损失以“-”号记）',
                               '其他收益', '营业利润', '营业外收入', '营业外支出',
                               '利润总额', '所得税费用', '实际上缴税费总额',
                               '其中：增值税', '所得税', '减免税总额',
                               '其中：增值税.1', '所得税.1', '其中：享受高新技术企业所得税减免',
                               '研发加计扣除所得税减免', '技术转让所得税减免', '应交增值税',
                               '本年应付职工薪酬（本年贷方累计发生额）', '年末资产总计',
                               '其中：流动资产合计', '非流动资产合计',
                               '其中：固定资产净值', '其中：无形资产',
                               '其中：累计折旧', '其中：本年折旧', '年末负债合计',
                               '其中：银行贷款额', '其中：当年获得银行贷款金额',
                               '年末所有者权益（股东权益）', '其中：实收资本（股本）',
                               '其中：企业上市融资股本', '其中：企业海外上市融资股本',
                               '对境外直接投资额', '本年完成固定资产投资额',
                               '规模以上工业企业及重点耗能企业综合能源消费量',
                               '企业当年是否获得风险投资',
                               '若是，请注明企业获得的风险投资的阶段是',
                               '当年获得创业风险投资机构的风险投资额',
                               '其中：管理和服务人员',
                               '其中：全职人员', '其中：本科毕业及以上人员',
                               '其中：外聘人员',
                               '1.人员人工费用', '2.直接投入费用',
                               '3.折旧费用与长期待摊费用', '4.无形资产摊销费用',
                               '5.设计费用', '6.装备调试费用与试验费用',
                               '7.委托外单位开展科技活动费用合计',
                               '其中：委托境内研究机构', '委托境内高等学校',
                               '委托境内企业', '委托境外机构', '8.其他费用',
                               '当年形成用于科技活动的固定资产',
                               '其中：仪器和设备', '期末机构数',
                               '机构研究开发人员合计', '其中：博士毕业',
                               '硕士毕业', '机构研究开发费用',
                               '其中：申请发明专利',
                               '其中：申请国内发明专利', '其中：申请欧美日专利',
                               '其中：申请PCT国际专利', '当年专利授权数',
                               '其中：授权发明专利', '其中：授权国内发明专利',
                               '其中：授权欧美日专利', '期末拥有有效专利数',
                               '其中：拥有境外授权专利', '其中：拥有欧美日专利',
                               '其中：拥有发明专利', '其中：已被实施的发明专利',
                               '其中：境外授权发明专利', '专利所有权转让及许可数',
                               '专利所有权转让及许可收入', '新产品产值',
                               '新产品销售收入', '其中：出口',
                               '期末境外营销服务机构数', '期末境外研究开发机构数',
                               '期末境外生产制造基地数', '当年在境外设立分支机构数量',
                               '发表科技论文', '期末拥有注册商标',
                               '其中：当年注册商标', '其中：境外注册商标',
                               '其中：当年境外注册商标', '拥有软件著作权',
                               '其中：当年获得软件著作权', '拥有集成电路布图',
                               '其中：当年获得集成电路布图', '拥有植物新品种',
                               '其中：当年获得植物新品种', '拥有国家一类新药品种',
                               '其中：当年获得国家一类新药证书',
                               '拥有国家一级中药保护品种',
                               '其中：当年获得国家一级中药保护品种证书',
                               '当年形成国际标准', '当年形成国家或行业标准',
                               '当年获得国家科技奖励', '认定登记的技术合同项数',
                               '认定登记的技术合同成交金额',
                               '使用来自政府部门的科技活动经费',
                               '技术改造经费支出', '购买境内技术经费支出',
                               '引进境外技术经费支出',
                               '引进境外技术的消化吸收经费支出',
                               '①美国：', '②欧盟：', '③日本：', '④其他：',
                               '①自主研发：', '②受让：', '③受赠：',
                               '④并购：', '⑤通过5年以上的独占许可：'
                               ]
            df = df.drop(columns=columns_to_drop)

            # 去掉列名后面的冒号
            # 除掉列名中的带圈序号
            # 除掉列名中的逗号
            df.rename(columns=lambda col: re.sub(r'[，：.:,（）()]', '', col), inplace=True)

            # 保存回一个新的CSV文件
            cleanDataTableName = 'clean_data.csv'
            df.to_csv(cleanDataTableName, index=False)

            # 4.将清洗后的数据导入到hive
            # 根据清洗过后的表头在hive建表
            importHive(cleanDataTableName)

            return render_template('upload.html', ok="上传成功！")

        except Exception as e:
            return render_template('upload.html', error=f"上传失败: {str(e)}")


# 分析各个地区的高企数量:ok
@app.route("/high_tech_count", methods=['GET'])
def get_high_tech_count():
    """
    获取各地区高企数量
    """
    # 1. 连接 Hive
    cursor, conn = connectHive()
    try:
        # 2. 查询数据
        query = """
        SELECT `行政区划代码`, COUNT(*)
        FROM cleanData
        GROUP BY `行政区划代码`
        """
        cursor.execute(query)

        # 3. 处理查询结果
        result = cursor.fetchall()
        high_tech_count = {}
        for row in result:
            code, count = row
            region_name = areas1.get(code[:4], "未知地区")  # 使用 areas1 获取地区名称
            high_tech_count[region_name] = count

        # 返回结果为 JSON 格式
        return jsonify({
            "status": "success",
            "data": high_tech_count
        })

    except Exception as e:
        # 处理异常
        return jsonify({
            "status": "error",
            "message": str(e)
        })

    finally:
        # 关闭连接
        try:
            closeHive(cursor, conn)
        except:
            pass


# 分析各地市高企质量：ok
@app.route("/enterpriseQuality")
def enterpriseQuality():
    # 获取所有数据使用K-means聚类算法将所有数据分类
    # 统计这三类里，各个地市的占比,曲线图表示
    # 需要的数据：科技活动人员、从业人员年平均人数、科技活动费用合计、净利润、当年申请专利数量

    # 1、连接hive
    cursor, conn = connectHive()
    # 2、获取所有的数据
    sql = "select `行政区划代码`,`净利润`,`从业人员年平均人数`,`科技活动费用合计`,`科技活动人员合计`,`当年专利申请数` from cleanData"
    cursor.execute(sql)
    datas = cursor.fetchall()
    closeHive(cursor,conn)
    # 3、将数据存到一个DataFrame里
    # 将查询结果转换为pandas DataFrame
    df = pd.DataFrame(datas,
                      columns=['行政区划代码', '净利润', '从业人员年平均人数', '科技活动费用合计', '科技活动人员合计','当年专利申请数'])
    # 将行政区划代码转换为地区名称
    df['地区名称'] = df['行政区划代码'].apply(lambda x: areas1.get(x, '未知'))
    # 删除行政区划代码列
    df.drop(columns=['行政区划代码'], inplace=True)
    # 增加一列科技活动人员在从业年平均人数的占比
    df['科技活动人员占比'] = df['科技活动人员合计'] / df['从业人员年平均人数']

    # 4、使用聚类算法进行分类
    features = df[['净利润', '从业人员年平均人数', '科技活动费用合计', '科技活动人员合计', '当年专利申请数']]
    # 使用KMeans进行聚类，这里假设我们想要分成3类
    kmeans = KMeans(n_clusters=4, random_state=42)
    df['聚类标签'] = kmeans.fit_predict(features)

    # 将聚类标签重新命名为低、中、高、超高
    label_mapping = {0: '低' , 1: '中' , 2: '高' , 3: '超高'}
    df['聚类标签'] = df['聚类标签'].map(label_mapping)
    print(df['聚类标签'])
    # 将数据转换为JSON格式
    result_json = df.to_dict(orient='records')
    # 5、统计每一类中各个地市的占比
    cluster_stats = df.groupby(['聚类标签' , '地区名称']).size().unstack(fill_value=0)
    cluster_stats_percent = cluster_stats.div(cluster_stats.sum(axis=1) , axis=0).round(2)
    # 将DataFrame转换为JSON格式的数据
    # 注意：这里我们只需要聚类标签和地区名称的占比信息
    result_json = cluster_stats_percent.reset_index().to_dict(orient='records')
    print(cluster_stats_percent)
    # 返回JSON格式数据
    return json.dumps(result_json, ensure_ascii=False)


# 预处理结果

@app.route("/getCleanData", methods=["POST", "GET"])
def getCleanData():
    # 对于 GET 请求，使用 query 参数；对于 POST 请求，使用 form 数据
    area = request.args.get('city') if request.method == 'GET' else request.form.get('city')

    if not area:
        area = 'all'  # 如果没有提供城市参数，默认获取所有数据

    print(f'city: {area}')

    try:
        cursor, conn = connectHive()  # 连接到 Hive 数据库

        if area != 'all':
            # 根据地区查找数据
            sql = 'SELECT * FROM cleanData WHERE `行政区划代码`=%s'
            cursor.execute(sql, (areas[area],))
        else:
            sql = 'SELECT * FROM cleanData'
            cursor.execute(sql)

        datas = cursor.fetchall()
        result = []

        for tuple_item in datas:
            tableId, account, dataStatus, administrativeRegionCode, industryCode2017, registeredCapital, listingStatus, \
                listingDate, isListedEntity, marketValue, industryOutputValue, revenue, mainBusinessIncome, \
                technologyIncome, technologyTransferIncome, technologyContractIncome, technologyConsultingIncome, \
                commissionedResearchIncome, productSalesIncome, highTechProductIncome, goodsSalesIncome, otherRevenue, \
                netProfit, employeesEndOfPeriod, returnedStudents, foreignResidents, foreignExperts, newEmployees, graduateStudents, \
                averageEmployees, postgraduateEmployees, doctorateEmployees, masterEmployees, bachelorEmployees, associateDegreeEmployees, \
                nonEmployedBeforeHigherEducation, technicians, seniorTechnicians, techniciansLevel2, seniorSkillsLevel3, intermediateSkillsLevel4, \
                juniorSkillsLevel5, middleManagementPersonnel, professionalTechnicalPersonnel, scienceAndTechnologyPersonnel, \
                scienceAndTechnologyExpenses, patentApplications = tuple_item
            dataItem = {
                'id': tableId,
                'account': account,
                'dataStatus': dataStatus,
                'administrativeRegionCode': administrativeRegionCode,
                'industryCode2017': industryCode2017,
                'registeredCapital': registeredCapital,
                'listingStatus': listingStatus,
                'listingDate': listingDate,
                'isListedEntity': isListedEntity,
                'marketValue': marketValue,
                'industryOutputValue': industryOutputValue,
                'revenue': revenue,
                'mainBusinessIncome': mainBusinessIncome,
                'technologyIncome': technologyIncome,
                'technologyTransferIncome': technologyTransferIncome,
                'technologyContractIncome': technologyContractIncome,
                'technologyConsultingIncome': technologyConsultingIncome,
                'commissionedResearchIncome': commissionedResearchIncome,
                'productSalesIncome': productSalesIncome,
                'highTechProductIncome': highTechProductIncome,
                'goodsSalesIncome': goodsSalesIncome,
                'otherRevenue': otherRevenue,
                'netProfit': netProfit,
                'employeesEndOfPeriod': employeesEndOfPeriod,
                'returnedStudents': returnedStudents,
                'foreignResidents': foreignResidents,
                'foreignExperts': foreignExperts,
                'newEmployees': newEmployees,
                'graduateStudents': graduateStudents,
                'averageEmployees': averageEmployees,
                'postgraduateEmployees': postgraduateEmployees,
                'doctorateEmployees': doctorateEmployees,
                'masterEmployees': masterEmployees,
                'bachelorEmployees': bachelorEmployees,
                'associateDegreeEmployees': associateDegreeEmployees,
                'nonEmployedBeforeHigherEducation': nonEmployedBeforeHigherEducation,
                'technicians': technicians,
                'seniorTechnicians': seniorTechnicians,
                'techniciansLevel2': techniciansLevel2,
                'seniorSkillsLevel3': seniorSkillsLevel3,
                'intermediateSkillsLevel4': intermediateSkillsLevel4,
                'juniorSkillsLevel5': juniorSkillsLevel5,
                'middleManagementPersonnel': middleManagementPersonnel,
                'professionalTechnicalPersonnel': professionalTechnicalPersonnel,
                'scienceAndTechnologyPersonnel': scienceAndTechnologyPersonnel,
                'scienceAndTechnologyExpenses': scienceAndTechnologyExpenses,
                'patentApplications': patentApplications

            }
            result.append(dataItem)

        if not result:
            return jsonify({"result": [], "status": 204, "message": "No data found."}), 204
        result = {"result": result, "status": 200}
        print(datas)
        data = json.dumps(result, ensure_ascii=False)

        return jsonify(result)
        #return render_template('yuchuli.html',data=result)
        #return jsonify({"result": result, "status": 200})  # 返回数据


    except Exception as e:
        print(f"Error occurred: {e}")
        return jsonify({"error": str(e), "status": 500}), 500

    finally:
        if 'cursor' in locals() and 'conn' in locals():
            closeHive(cursor, conn)

# 展示预处理结果路由
@app.route("/yuchuli")
def yuchuli():
    print("Rendering yuchuli.html")  # 调试输出，检查是否进入函数
    return render_template('yuchuli.html')  # 渲染 yuchuli.html 页面


# 分析各地市高企的产业分布情况
@app.route("/chanYe/<string:area>", methods=["GET"])
def chanYe(area):
    # 1需要的数据有：cleanData表里的行政区化代码、行业代码、code表里的产业名
    cursor, conn = connectHive()
    sql = 'SELECT c.`行业名`,COUNT(*) AS `行业数量` FROM cleanData AS cd JOIN code AS c ON cd.`行业代码2017` = c.`行业代码` where cd.`行政区划代码`=%s GROUP BY c.`行业名`'
    cursor.execute(sql, (areas[area],))
    datas = cursor.fetchall()
    print(datas)
    # 使用 for 循环遍历列表中的每个元组
    results = []
    # 使用 for 循环遍历列表中的每个元组
    for tuple_item in datas:
        # 解包元组，将元素分别赋值给变量
        chanYeName, totalNum = tuple_item
        # 创建一部字典来存储元组数据
        dataItem = {
            'chanYeName': chanYeName,
            'totalNum': totalNum
        }
        # 将字典添加到 json_data 列表中
        results.append(dataItem)

        # 打印每个元组的内容
    a_json = {"result": results, "status": 200}
    print(datas)
    closeHive(cursor, conn)
    return json.dumps(a_json, ensure_ascii=False)


# 各地市高企研发总投入分析:ok
@app.route("/getTotalInvestment")
def getTotalInvestment():
    # 需要的数据：行政区划代码、科技活动费总计
    # 按地市分类，统计该地市所有高企的研发费用
    # 返回的数据是：地市，研发总投入

    # 1、连接hive
    cursor, conn = connectHive()
    # 2、查询数据
    cursor.execute('select `行政区划代码`,sum(`科技活动费用合计`) from cleanData group by `行政区划代码`')

    datas = cursor.fetchall()
    print(datas)
    # 3、将数据封装成json格式
    results = []
    # 使用 for 循环遍历列表中的每个元组
    for tuple_item in datas:
        # 解包元组，将元素分别赋值给变量
        area, totalInvestment = tuple_item
        # 创建一部字典来存储元组数据
        dataItem = {
            'area': areas1[area],
            'totalInvestment': totalInvestment
        }
        # 将字典添加到 json_data 列表中
        results.append(dataItem)

        # 打印每个元组的内容
    a_json = {"result": results, "status": 200}
    print(datas)
    closeHive(cursor, conn)
    # 4、返回封装后的数据
    return json.dumps(a_json, ensure_ascii=False)


# 各地市上市企业的比例：饼图 ok
@app.route("/getMarketRate")
def getMarketRate():
    # 各地市的上市企业数量与总企业数量的比值
    # 1.统计各地市上市企业的数量
    # 连接hive
    cursor, conn = connectHive()
    cursor.execute(
        'select `行政区划代码`,count(*) from cleanData where `上市及新三板、四板挂牌情况`!="0" group by `行政区划代码`')
    datasOnMarket = cursor.fetchall()
    print(datasOnMarket)
    closeHive(cursor, conn)
    # 2.统计各地市的总企业数量
    cursor, conn = connectHive()
    cursor.execute('select `行政区划代码`,count(*) from cleanData group by `行政区划代码`')
    datasTotal = cursor.fetchall()
    print(datasTotal)
    closeHive(cursor, conn)
    # 3.创建新的字典{地区，比值}
    # 将数据封装成json格式
    results = []
    # 使用 for 循环遍历列表中的每个元组
    for tuple_item1 in datasTotal:
        # 解包元组，将元素分别赋值给变量
        area1, totalNum = tuple_item1
        flat = True
        for tuple_item2 in datasOnMarket:
            area2, onMarketNum = tuple_item2
            if area1 in tuple_item2:
                # 创建一部字典来存储元组数据
                dataItem = {
                    'area': areas1[area1],
                    'marketRate': onMarketNum / totalNum
                }
                flat = False
                results.append(dataItem)
                break
                # 将字典添加到 json_data 列表中
        if flat:
            dataItem = {
                'area': areas1[area1],
                'marketRate': 0
            }
            results.append(dataItem)
        # 打印每个元组的内容
    a_json = {"result": results, "status": 200}
    print(a_json)
    # print(datas)
    # 4、返回封装后的数据
    return json.dumps(a_json, ensure_ascii=False)


# 各地市高企的盈亏情况(即各地市净利润为正的高企数量与该地市的总高企数量之比)：曲线图 ok
@app.route("/getProfitRate")
def getProfitRate():
    # 1.统计各地市净利润为正的高企数量
    # 连接hive
    cursor, conn = connectHive()
    cursor.execute(
        'select `行政区划代码`,count(*) from cleanData where `净利润`>0 group by `行政区划代码`')
    positiveProfitNum = cursor.fetchall()
    print(positiveProfitNum)
    closeHive(cursor, conn)
    # 2.统计各地市的总企业数量
    cursor, conn = connectHive()
    cursor.execute('select `行政区划代码`,count(*) from cleanData group by `行政区划代码`')
    TotalNum = cursor.fetchall()
    print(TotalNum)
    closeHive(cursor, conn)
    results = []
    # 使用 for 循环遍历列表中的每个元组
    for tuple_item1 in TotalNum:
        # 解包元组，将元素分别赋值给变量
        area1, totalNum = tuple_item1
        flat = True
        for tuple_item2 in positiveProfitNum:
            area2, posPNum = tuple_item2
            if area1 in tuple_item2:
                # 创建一部字典来存储元组数据
                dataItem = {
                    'area': areas1[area1],
                    'positiveProfit': posPNum / totalNum
                }
                flat = False
                results.append(dataItem)
                break
                # 将字典添加到 json_data 列表中
        if flat:
            dataItem = {
                'area': areas1[area1],
                'positiveProfit': 0
            }
            results.append(dataItem)
        # 打印每个元组的内容
    a_json = {"result": results, "status": 200}
    print(a_json)
    # print(datas)
    # 4、返回封装后的数据
    return json.dumps(a_json, ensure_ascii=False)


# 各地高企从业人员分布情况:饼图:ok
@app.route("/get_employment_data", methods=['GET'])
def get_employment_data():
    # 连接到Hive
    cursor, conn = connectHive()

    # 执行查询语句
    query = """
    SELECT `行政区划代码`, SUM(`从业人员期末人数`)
    FROM cleanData
    GROUP BY `行政区划代码`
    """
    cursor.execute(query)

    # 获取查询结果
    results = cursor.fetchall()

    # 关闭Hive连接
    closeHive(cursor, conn)

    # 将查询结果转换为pandas DataFrame
    df = pd.DataFrame(results, columns=['行政区划代码', '从业人员期末人数'])

    # 将行政区划代码转换为地区名称
    df['地区名称'] = df['行政区划代码'].apply(lambda x: areas1.get(x, '未知'))

    # 删除行政区划代码列
    df.drop(columns=['行政区划代码'], inplace=True)

    # 将数据转换为JSON格式
    result_json = df.to_dict(orient='records')

    # 返回JSON格式数据
    return json.dumps(result_json, ensure_ascii=False)


# 企业注册资金分析
# 低注册资金： 注册资金 < 50000 千元
# 中等注册资金： 50000 千元 <= 注册资金 < 200000 千元
# 高注册资金： 注册资金 >= 200000 千元
@app.route("/get_registration_data_by_category", methods=['GET'])
def get_registration_data_by_category():
    # 连接到Hive
    cursor, conn = connectHive()

    # 执行查询语句
    query = """
    SELECT 
        `行政区划代码`,
        CASE 
            WHEN `注册资金` < 50000 THEN '低注册资金'
            WHEN `注册资金` BETWEEN 50000 AND 200000 THEN '中等注册资金'
            ELSE '高注册资金'
        END AS `注册资金分类`,
        COUNT(*) AS `企业数量`
    FROM cleanData
    GROUP BY `行政区划代码`, 
             CASE 
                 WHEN `注册资金` < 50000 THEN '低注册资金'
                 WHEN `注册资金` BETWEEN 50000 AND 200000 THEN '中等注册资金'
                 ELSE '高注册资金'
             END
    """
    cursor.execute(query)

    # 获取查询结果
    results = cursor.fetchall()

    # 关闭Hive连接
    closeHive(cursor, conn)

    # 将查询结果转换为pandas DataFrame
    df = pd.DataFrame(results, columns=['行政区划代码', '注册资金分类', '企业数量'])

    # 将行政区划代码转换为地区名称
    df['地区名称'] = df['行政区划代码'].apply(lambda x: areas1.get(x, '未知'))

    # 构造输出结果
    grouped_data = df.pivot_table(
        index='地区名称',
        columns='注册资金分类',
        values='企业数量',
        aggfunc='sum',
        fill_value=0
    ).reset_index()

    # 重命名列名，匹配期望输出
    grouped_data.rename(
        columns={
            '低注册资金': '低注册资金企业数量',
            '中等注册资金': '中等注册资金企业数量',
            '高注册资金': '高注册资金企业数量'
        },
        inplace=True
    )

    # 转换为期望的 JSON 格式
    result_json = []
    for row in grouped_data.to_dict(orient="records"):
        result_json.append(OrderedDict([
            ("地区名称", row["地区名称"]),
            ("低注册资金企业数量", row.get("低注册资金企业数量", 0)),
            ("中等注册资金企业数量", row.get("中等注册资金企业数量", 0)),
            ("高注册资金企业数量", row.get("高注册资金企业数量", 0)),
        ]))

    # 返回 JSON 响应
    return json.dumps(result_json, ensure_ascii=False)

#分析企业各项营业收入的占比：饼图
@app.route('/get_income_data', methods=['GET'])
def get_income_data():
    try:
        # 1. 连接到 Hive 数据库
        cursor, conn = None, None
        try:
            print("Connecting to Hive database...")
            cursor, conn = connectHive()
            print("Connected to Hive database.")
        except Exception as conn_error:
            print("Database connection failed:", conn_error)
            raise RuntimeError(f"Database connection failed: {conn_error}")

        # 2. 查询数据
        try:
            query = """
            SELECT
                `行政区划代码` AS `region_code`,
                SUM(`营业收入`) AS `总收入`,
                SUM(`其中主营业务收入`) AS `主营业务收入`,
                SUM(`其中技术收入`) AS `技术收入`,
                SUM(`其中技术转让收入`) AS `技术转让收入`,
                SUM(`技术承包收入`) AS `技术承包收入`,
                SUM(`技术咨询与服务收入`) AS `技术咨询与服务收入`,
                SUM(`接受委托研究开发收入`) AS `接收委托研究开发收入`,
                SUM(`产品销售收入`) AS `产品销售收入`,
                SUM(`其他营业收入`) AS `其他营业收入`
            FROM cleanData
            GROUP BY `行政区划代码`
            """
            print("Executing query...")
            cursor.execute(query)
            print("Query executed successfully.")

            columns = [desc[0] for desc in cursor.description]
            print("Columns fetched from query:", columns)
            rows = cursor.fetchall()
            print("Rows fetched:", rows)

            data = pd.DataFrame(rows, columns=columns)
            print("Data loaded into DataFrame.")
        except Exception as query_error:
            print("Error during query execution:", query_error)
            raise RuntimeError(f"Data query failed: {query_error}")

        # 3. 数据处理
        try:
            print("Processing data...")
            income_columns = [
                "主营业务收入", "技术收入", "技术转让收入", "技术承包收入",
                "技术咨询与服务收入", "接收委托研究开发收入", "产品销售收入", "其他营业收入"
            ]

            # 使用 Decimal 进行精确计算
            for col in income_columns:
                data[f"{col}_占比"] = data.apply(
                    lambda row: (Decimal(row[col]) / Decimal(row["总收入"]) * 100)
                    if row["总收入"] > 0 else 0, axis=1
                )
            print("Data processing completed.")

            # 替换行政区代码为中文名称
            if 'region_code' in data.columns:  # 检查是否存在该列
                data['行政区划名称'] = data['region_code'].map(areas1).fillna("未知地区")
            else:
                raise RuntimeError("数据中未找到列 'region_code'")
        except Exception as processing_error:
            print("Error during data processing:", processing_error)
            raise RuntimeError(f"Data processing failed: {processing_error}")

        # 4. 关闭连接
        try:
            closeHive(cursor, conn)
            print("Hive connection closed.")
        except Exception as close_error:
            print("Error closing Hive connection:", close_error)

        # 5. 转换结果为 JSON 并返回
        print("Converting data to JSON response.")
        result = data.to_dict(orient="records")
        return jsonify({
            "status": "success",
            "data": result
        })

    except RuntimeError as e:
        print("Runtime error occurred:", e)
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

    except Exception as e:
        print("Unexpected error occurred:", e)
        return jsonify({
            "status": "error",
            "message": "Unexpected error occurred",
            "details": str(e)
        }), 500

# 连接hive的函数
def connectHive():
    conn = connect(host="121.36.46.125", port=10000, user="root", password="Hadoop123456",
                   auth_mechanism="PLAIN")
    print("1")
    # 连接后，获取游标，标志指示数据库的当前记录
    cursor = conn.cursor()
    return cursor, conn


# 关闭hive的函数
def closeHive(cursor, conn):
    cursor.close()
    conn.close()


# 测试可不可根据表头信息自动建表，字段太多了，无法手动建表
def importHive(tableName):
    # 连接hive2的服务器
    # 1.读取属性表
    result = ""
    df = pd.read_csv(tableName)

    # 判断数据类型
    # 初始化一部字典来存储列名和数据类型的映射
    column_types = {}
    # 遍历每列，推断数据类型
    print(df.columns)
    for column in df.columns:
        print(column)
        # 检查空值
        if df[column].isnull().all():
            # 如果整列都是空值，可以指定一个默认类型（如STRING）或跳过该列
            column_types[column] = 'STRING'
        else:
            # 尝试推断数据类型
            unique_values = df[column].unique()
            print(unique_values)
            if all(isinstance(value, int) for value in unique_values if pd.notnull(value)):
                # 如果所有非空值都是整数，则指定为INT类型
                column_types[column] = 'INT'
            elif all(isinstance(value, float) for value in unique_values if pd.notnull(value)):
                # 如果所有非空值都是浮点数，则指定为FLOAT类型
                column_types[column] = 'FLOAT'
            elif all(isinstance(value, str) for value in unique_values if pd.notnull(value)):
                # 如果所有非空值都是字符串，则指定为STRING类型
                # 注意：这里可能需要进一步的检查来确定是否为日期或其他特定类型的字符串
                column_types[column] = 'STRING'
            else:
                column_types[column] = 'STRING'

    column_types['从业人员期末人数'] = 'INT'
    column_types['其中留学归国人员'] = 'INT'
    column_types['其中外籍常驻人员'] = 'INT'
    column_types['其中引进外籍专家'] = 'INT'
    column_types['其中当年新增从业人员'] = 'INT'
    column_types['其中吸纳高校应届毕业生'] = 'INT'
    column_types['从业人员年平均人数'] = 'INT'
    column_types['具有研究生学历位人员'] = 'INT'
    column_types['其中博士'] = 'INT'
    column_types['其中硕士'] = 'INT'
    column_types['具有大学本科学历位人员'] = 'INT'
    column_types['具有大学专科学历人员'] = 'INT'
    column_types['接受高等教育前为非就业地户籍人员'] = 'INT'
    column_types['技能人员'] = 'INT'
    column_types['其中高级技师国家职业资格一级'] = 'INT'
    column_types['技师国家职业资格二级'] = 'INT'
    column_types['高级技能人员国家职业资格三级'] = 'INT'
    column_types['中级技能人员国家职业资格四级'] = 'INT'
    column_types['初级技能人员国家职业资格五级'] = 'INT'
    column_types['中层及以上管理人员'] = 'INT'
    column_types['专业技术人员'] = 'INT'
    column_types['科技活动人员合计'] = 'INT'
    column_types['当年专利申请数'] = 'INT'

    print(column_types)
    flat = False
    for (key, value) in column_types.items():
        if flat:
            result += ','
        result += f"`{key}` {value}"
        flat = True
    print(result)

    # 填充空值
    df.fillna(-2147483640, inplace=True)
    # 新建一个CSV文件，不读取clean_data.csv文件的表头，为了将清洗后的数据导入到hive时不将表头也一起导入
    df.columns = range(df.shape[1])
    # 然后将DataFrame（不包含原表头）写回到一个新的CSV文件中
    # 注意：我们使用header=False来避免写入新的列名（表头）
    new_csv_filename = 'cleanData.csv'
    df.to_csv(new_csv_filename, index=False, header=False)

    conn = connect(host="121.36.46.125", port=10000, user="root", password="Hadoop123456",
                   auth_mechanism="PLAIN")
    print("1")
    # 连接后，获取游标，标志指示数据库的当前记录
    cursor = conn.cursor()
    table_name = 'cleanData'
    create_table_sql = f"CREATE TABLE IF NOT EXISTS {table_name} ({result}) ROW FORMAT DELIMITED FIELDS TERMINATED BY ',' STORED AS TEXTFILE"
    print(create_table_sql)
    cursor.execute(create_table_sql)

    # 3.将文件上传到HDFS
    # 创建HDFS客户端

    client = InsecureClient(HDFS_URL, user=HDFS_USER)
    # 上传文件到HDFS
    hdfs_path = HDFS_UPLOAD_DIR + tableName
    print(hdfs_path)
    client.upload(hdfs_path, new_csv_filename)
    cursor = conn.cursor()
    cursor.execute(f"LOAD DATA INPATH '{hdfs_path}' INTO TABLE {table_name} ")
    print("2")


# 测试连接hive
@app.route("/linkHive")
def link():
    # 连接hive2的服务器
    conn = connect(host="121.36.46.125", port=10000, user="root", password="Hadoop123456", auth_mechanism="PLAIN")
    # 连接后，获取游标，标志指示数据库的当前记录
    cursor = conn.cursor()
    # 游标帮助我们执持HQL语句
    # cursor.execute('create table test1(id int,name string)')
    # 查询的结果在cursor.fetchall方法中
    # cursor.execute("select * from default.mytest")
    cursor.execute('select `行政区划代码`,count(*) from cleanData group by `行政区划代码`')
    results = cursor.fetchall()
    a_json = {"result": results, "status": 200}
    print(results)
    return json.dumps(a_json, ensure_ascii=False)
    # return "ok"
    # return json.dumps({"result": "网络连接错误", "status": 404}, ensure_ascii=False)


if __name__ == '__main__':
    app.run(debug=True, port=5200)
